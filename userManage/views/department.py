from openpyxl import load_workbook

from userManage import models
from django.shortcuts import render, redirect, HttpResponse
from userManage.utils.pagination import Pagination


def depart_list(request):
    """部门列表"""
    depart_data = models.Department.objects.all()

    # 分页
    page_object = Pagination(request, depart_data)
    context = {
        "depart_data": page_object.page_queryset,
        "page_string": page_object.html(),
    }
    return render(request, 'depart_list.html', context)


def depart_add(request):
    """添加页面"""
    if request.method == 'GET':
        return render(request, 'depart_add.html')
    depart_name = request.POST.get('title')
    models.Department.objects.create(department_title=depart_name)
    return redirect('/depart/list/')


def depart_delete(request):
    """删除部门"""
    depart_id = request.GET.get('id')
    models.Department.objects.filter(id=depart_id).delete()
    return redirect('/depart/list/')


def depart_edit(request, id):
    """编辑部门"""
    if request.method == 'GET':
        depart = models.Department.objects.filter(id=id).first()
        return render(request, "depart_edit.html", {'depart': depart})
    else:
        # 获取用户输入的新部门名
        name = request.POST.get("title")
        # 根据id找到数据库中的数据并进行更新
        models.Department.objects.filter(id=id).update(department_title=name)
        return redirect('/depart/list/')


def depart_multi(request):
    """批量上传(Excel文件)"""
    # 获取用户上传的文件对象
    file_object = request.FILES.get("depart-excel")
    # print(file_object)
    # 将文件对象传递给openpyxl，读取文件内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 循环获取每一行数据, 从第二行开始
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value   # 每一行的第一列
        # 添加到数据库中
        # 判断数据库中是否已经存在该数据
        exists = models.Department.objects.filter(department_title=text).exists()
        if not exists:
            models.Department.objects.create(department_title=text)
    return redirect("/depart/list/")
